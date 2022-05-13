from openpyxl import Workbook
from openpyxl import load_workbook

class Escrever:
    def __init__(self, nome_arquivo):

        self.nome_planilha = nome_arquivo
        # nome_planilha = r'C:\Users\bruno.silva\Documents\pythonContabil\ATIVA-SYSTEM\planilhas\Contas Horus teste.xlsx'
        print(self.nome_planilha)

        if "')" in self.nome_planilha:
            self.nome_planilha = self.nome_planilha.replace("')", '')

        # nome_planilha = nome_planilha.split('/')
        # self.nome_planilha = nome_planilha[-1]
        self.wb = Workbook()
        self.wb.save(f'{self.nome_planilha}.xlsx')
        self.wa = load_workbook(f'{self.nome_planilha}.xlsx')
        self.ws = self.wa.active

    def escrevendo(self, lista):
        for linha in lista:
            self.ws.append(linha)
        self.ws.insert_rows(1)
        self.ws['A1'] = "Classificação"
        self.ws['B1'] = "Código Reduzido"
        self.ws['C1'] = "Nome da Conta "
        self.ws['D1'] = "Natureza da Operação"
        self.ws['E1'] = "Tipo da Conta"
        self.wa.save(f'{self.nome_planilha}.xlsx')
    #
    # def escrevendobanco(self, listab, listac):
    #     self.wc = Workbook()
    #     self.wc.save('sistemadominio.xlsx')
    #     self.wd = load_workbook('sistemadominio.xlsx')
    #     self.we = self.wd.active
    #
    #     for linha in listab:
    #         self.we.append(linha)
    #     for linha in listac:
    #         self.we.append(linha)
    #     self.we.insert_rows(1)
    #     self.we['A1'] = "Data"
    #     self.we['B1'] = "Numero"
    #     self.we['C1'] = "Valor Credito"
    #     self.we['D1'] = "Valor Debito"
    #     self.wd.save('sistemadominio.xlsx')



    def salvar(self):
        self.wb.save('testefinal.xlsx')

if __name__ == '__main__':
    e = Escrever(r'file:///C:/Users/bruno.silva/Documents/pythonContabil/ATIVA-SYSTEM/teste4')
    # e.escrevendo([(1,2,3,4,5,6),(1,2,3,4,5,6),(1,2,3,4,5,6),(1,2,3,4,5,6),(1,2,3,4,5,6),(1,2,3,4,5,6),(1,2,3,4,5,6)])