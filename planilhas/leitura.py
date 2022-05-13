from planilhas.escrevendo import Escrever
from openpyxl import load_workbook
import datetime

class Ler_planilha:
    def __init__(self, arquivo):
        self.wb = load_workbook(rf'{arquivo}')
        self.celula = self.wb['Contas']

    def ler(self):
        print('iniciou')
        self.listaH = []
        self.listaT = []
        for row in range(6, self.celula.max_row + 1):
            self.codigigo = self.celula[f'A{row}'].value
            self.classificacao = self.celula[f'H{row}'].value
            self.nome1 = self.celula[f'L{row}'].value
            self.nome2 = self.celula[f'M{row}'].value
            self.nome3 = self.celula[f'N{row}'].value
            self.nome4 = self.celula[f'O{row}'].value
            self.nome5 = self.celula[f'P{row}'].value
            self.nome6 = self.celula[f'Q{row}'].value
            self.nome7 = self.celula[f'R{row}'].value
            self.nome8 = self.celula[f'S{row}'].value
            self.nome9 = self.celula[f'T{row}'].value
            # self.valorX = self.celula[f'X{row}'].value
            # self.valorD = self.celula[f'D{row}'].value
            # self.valorJ = self.celula[f'J{row}'].value
            # self.valorL = self.celula[f'L{row}'].value
            # self.valorO = self.celula[f'O{row}'].value
            # self.valorS = self.celula[f'S{row}'].value
            print(self.codigigo, self.classificacao)
            nome = f'{self.nome1} {self.nome2} {self.nome3} {self.nome4} {self.nome5}' \
                   f' {self.nome6} {self.nome7} {self.nome8} {self.nome9}'
            nome = nome.replace('None', '')
            print(nome)
            # self.listaH.append()

            # print(type(self.valorS))
            # print(type(self.valorO))
            # if type(self.valorS) == str:
            #     if "-" in self.valorS:
            #         self.valorS = self.valorS.replace("-", '')
            #         self.valorS = float(self.valorS)
            #         print(type(self.valorS))
            #
            #     else:
            #         pass

            # if "-" in self.valorO:
            #     print('existe numero engativo')
            # else:
            #     pass
            # self.valorO = float(self.valorO)
            # self.valorS = float(self.valorS)
            # if type(self.valorO) and type(self.valorS) == int:
            #     if self.valorO == self.valorS:
            #         resultado = self.valorO - self.valorS
            #         # print(f'{resultado:.2f}')
            #     else:
            #         resultado = self.valorO - self.valorS
            #         cel = self.celula[f'O{row}']
            #         # print(f'{resultado:.2f}')
            #         # print(f'Ouve divergencia na coluna O')
            # else:
            #     if type(self.valorO) == str:
            #         self.valorO = self.valorO
            #         self.valorO = self.valorO.replace("-", '')
            #     elif type(self.valorS) == str:
            #         self.valorS = self.valorS
            #         self.valorS = self.valorS.replace("-", '')
            # if self.valorO == None:
            #     pass
            # else:
            #     self.valorO = float(self.valorO)
            #
            # if self.valorS == None:
            #     pass
            # else:
            #     self.valorS = float(self.valorS)


        #
        #     historico = f"{self.valorD} {self.valorJ} {self.valorL}"
            conjunto = (self.classificacao, self.codigigo, nome)
            testo_info = f'classificação:{self.classificacao}, Codigo:{self.codigigo}, Descrição:{nome}'
            self.listaH.append(conjunto)
            self.listaT.append(testo_info)
        # print(self.listaH)
        #
        # Escrever().escrevendo(self.listaH)
        return self.listaH, self.listaT


if __name__ == '__main__':
    lp = Ler_planilha(r"C:\Users\bruno.silva\Documents\pythonContabil\ATIVA-SYSTEM\planilhas\Contas Horus.xlsx")
    lp.ler()
    # lp.escrever()
